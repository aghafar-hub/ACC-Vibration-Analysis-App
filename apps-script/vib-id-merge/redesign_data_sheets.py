import openpyxl, csv, re
from collections import Counter, defaultdict

XLSX = '/root/.claude/uploads/26301407-2e64-5285-b520-2fa1b4981fda/a8a11756-Vibration__Condition_Monitoring____ACC__V2.xlsx'
VIB_MAP_CSV = '/home/user/acc-vibration-analysis-app/apps-script/vib-id-merge/vib-point-map-import.csv'
OUT_DIR = '/home/user/acc-vibration-analysis-app/apps-script/vib-id-merge/'

def normalize(s):
    if s is None:
        return ''
    s = str(s).replace('\n', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    s = (s.replace('ouboard', 'outboard').replace('outbord', 'outboard'))
    return s.lower()

vib_lookup = {}
with open(VIB_MAP_CSV, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        eid, family, vib_id = row['Equipment ID'].strip(), row['Family'].strip().upper(), row['VIB ID']
        desc = row['Point Description']
        for part in [desc] + desc.split(';'):
            key = (eid, family, normalize(part))
            vib_lookup.setdefault(key, vib_id)

poscode_candidates = defaultdict(set)
with open(VIB_MAP_CSV, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        key = (row['Equipment ID'].strip(), row['Family'].strip().upper(), row['Position Code'].strip().upper())
        poscode_candidates[key].add(row['VIB ID'])
poscode_lookup = {k: next(iter(v)) for k, v in poscode_candidates.items() if len(v) == 1}

def derive_position_code(text):
    t = normalize(text)
    if not t:
        return None
    suffix = 'NDE' if ('outboard' in t or re.search(r'\bnde\b', t)) else ('DE' if ('inboard' in t or re.search(r'\bde\b', t)) else '')
    m = re.search(r'shaft\s*#?\s*(\d)', t)
    if m:
        return f'S{m.group(1)}{suffix}'
    if 'motor' in t:
        return f'M{suffix}'
    if 'compressor' in t or 'crusher' in t:
        return f'C{suffix}'
    if re.match(r'^bl\b', t):
        return f'BL{suffix}'
    if 'bearing' in t:
        return f'BR{suffix}'
    if 'fan' in t and suffix:
        return f'FN{suffix}'
    if 'gear box' in t or 'gearbox' in t or t == 'gb':
        return 'GB'
    m2 = re.search(r'rim\s*gear\s*(\d+)', t)
    if m2:
        return f'RG{m2.group(1)}'
    return None

ID_CORRECTIONS = {
    '645.BL580': '465.BL580',
    '645.BL630': '465.BL630',
    '645.BL635': '465.BL635',
}

wb = openpyxl.load_workbook(XLSX, data_only=True)

def fmt_date(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v)

def match(eid, family, point, extra_desc=None):
    key = (eid, family, normalize(extra_desc if extra_desc is not None else point))
    vib_id = vib_lookup.get(key, '')
    matched_by = 'text' if vib_id else None
    if not vib_id:
        pos = derive_position_code(point)
        if pos:
            vib_id = poscode_lookup.get((eid, family, pos), '')
            if vib_id:
                matched_by = 'poscode'
    return vib_id, matched_by

# ---------- RMS DATA ----------
ws = wb['📥 RMS DATA']
rms_out = []
rms_unmatched = []
rms_stats = Counter()
for r in range(4, ws.max_row + 1):
    eid = ws.cell(row=r, column=3).value
    point = ws.cell(row=r, column=4).value
    if not eid and not point:
        continue
    eid = ID_CORRECTIONS.get(str(eid).strip(), str(eid).strip()) if eid else ''
    point = point if point else ''
    date = ws.cell(row=r, column=5).value
    axial = ws.cell(row=r, column=6).value
    gear = ws.cell(row=r, column=7).value
    horizontal = ws.cell(row=r, column=8).value
    vertical = ws.cell(row=r, column=9).value
    vib_id, matched_by = match(eid, 'RMS', point)
    rms_stats['matched-' + matched_by if vib_id else 'unmatched'] += 1
    # column order per request: Horizontal, Vertical, Axial, Gear
    rms_out.append([vib_id, eid, ('' if vib_id else point), fmt_date(date), horizontal, vertical, axial, gear])
    if not vib_id:
        rms_unmatched.append([eid, point, fmt_date(date), r])

print('RMS DATA rows processed:', len(rms_out), dict(rms_stats))

# ---------- SPM DATA (split into SPM + GS) ----------
ws = wb['📥 SPM DATA']
spm_out = []
gs_out = []
spm_unmatched = []
spm_stats = Counter()
gs_count = 0
for r in range(4, ws.max_row + 1):
    eid = ws.cell(row=r, column=3).value
    point = ws.cell(row=r, column=4).value
    if not eid and not point:
        continue
    eid = ID_CORRECTIONS.get(str(eid).strip(), str(eid).strip()) if eid else ''
    point = point if point else ''
    type_val = ws.cell(row=r, column=5).value
    date = ws.cell(row=r, column=6).value
    hdm = ws.cell(row=r, column=7).value
    hdc = ws.cell(row=r, column=8).value
    gs = ws.cell(row=r, column=9).value

    type_str = str(type_val).strip() if type_val is not None else ''
    desc = f'{point} ({type_str})' if type_str else point
    vib_id, matched_by = match(eid, 'SPM', point, extra_desc=desc)
    spm_stats['matched-' + matched_by if vib_id else 'unmatched'] += 1
    spm_out.append([vib_id, eid, ('' if vib_id else point), fmt_date(date), hdm, hdc])
    if not vib_id:
        spm_unmatched.append([eid, point, type_str, fmt_date(date), r])

    if gs is not None and gs != '':
        gs_out.append([vib_id, eid, ('' if vib_id else point), fmt_date(date), gs])
        gs_count += 1

print('SPM DATA rows processed:', len(spm_out), dict(spm_stats))
print('GS rows extracted:', gs_count)

# ---------- write outputs ----------
def write_csv(path, header, rows, seq=True):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(header)
        for i, row in enumerate(rows, start=1):
            w.writerow(([i] if seq else []) + row)

write_csv(OUT_DIR + 'rms-data-redesigned.csv',
          ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Date', 'Horizontal (mm/s)', 'Vertical (mm/s)', 'Axial (mm/s)', 'Gear (mm/s)'],
          rms_out)
write_csv(OUT_DIR + 'spm-data-redesigned.csv',
          ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Date', 'HDm (dBsv)', 'HDc (dBsv)'],
          spm_out)
write_csv(OUT_DIR + 'gs-data.csv',
          ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Date', 'Gs'],
          gs_out)

with open(OUT_DIR + 'rms-data-unmatched.csv', 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['Equipment ID', 'Asset ID (point)', 'Date', 'Original row #'])
    w.writerows(rms_unmatched)
with open(OUT_DIR + 'spm-data-unmatched.csv', 'w', newline='', encoding='utf-8') as f:
    w = csv.writer(f)
    w.writerow(['Equipment ID', 'Asset ID (point)', 'Type', 'Date', 'Original row #'])
    w.writerows(spm_unmatched)

print()
print('Unique unmatched RMS (equipmentId, point) pairs:', len(set((u[0], u[1]) for u in rms_unmatched)))
print('Unique unmatched SPM (equipmentId, point, type) pairs:', len(set((u[0], u[1], u[2]) for u in spm_unmatched)))
