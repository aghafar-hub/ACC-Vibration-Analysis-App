import openpyxl, csv, re
from collections import Counter, defaultdict

XLSX = '/root/.claude/uploads/26301407-2e64-5285-b520-2fa1b4981fda/a8a11756-Vibration__Condition_Monitoring____ACC__V2.xlsx'
VIB_MAP_CSV = '/home/user/acc-vibration-analysis-app/apps-script/vib-id-merge/vib-point-map-import.csv'
OUT_DIR = '/home/user/acc-vibration-analysis-app/apps-script/vib-id-merge/'

RMS_SLOTS = 3   # covers 5143/5157 (VIB,month) combos exactly; the rare 14 with more go to an overflow file
SPM_SLOTS = 2   # covers all SPM/GS (VIB,month) combos seen in the data

def normalize(s):
    if s is None:
        return ''
    s = str(s).replace('\n', ' ')
    s = re.sub(r'\s+', ' ', s).strip()
    s = (s.replace('ouboard', 'outboard').replace('outbord', 'outboard'))
    return s.lower()

vib_lookup = {}
with open(VIB_MAP_CSV, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        eid, family, vib_id = row['Equipment ID'].strip(), row['Family'].strip().upper(), row['VIB ID']
        desc = row['Point Description']
        for part in [desc] + desc.split(';'):
            key = (eid, family, normalize(part))
            vib_lookup.setdefault(key, vib_id)

poscode_candidates = defaultdict(set)
with open(VIB_MAP_CSV, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        key = (row['Equipment ID'].strip(), row['Family'].strip().upper(), row['Position Code'].strip().upper())
        poscode_candidates[key].add(row['VIB ID'])
poscode_lookup = {k: next(iter(v)) for k, v in poscode_candidates.items() if len(v) == 1}

def derive_position_code(text):
    t = normalize(text)
    if not t:
        return None
    suffix = 'NDE' if ('outboard' in t or re.search(r'\bnde\b', t)) else ('DE' if ('inboard' in t or re.search(r'\bde\b', t)) else '')
    m = re.search(r'shaft\s*#?\s*(\d)', t)
    if m:
        return f'S{m.group(1)}{suffix}'
    if 'motor' in t:
        return f'M{suffix}'
    if 'compressor' in t or 'crusher' in t:
        return f'C{suffix}'
    if re.match(r'^bl\b', t):
        return f'BL{suffix}'
    if 'bearing' in t:
        return f'BR{suffix}'
    if 'fan' in t and suffix:
        return f'FN{suffix}'
    if 'gear box' in t or 'gearbox' in t or t == 'gb':
        return 'GB'
    m2 = re.search(r'rim\s*gear\s*(\d+)', t)
    if m2:
        return f'RG{m2.group(1)}'
    return None

ID_CORRECTIONS = {
    '645.BL580': '465.BL580',
    '645.BL630': '465.BL630',
    '645.BL635': '465.BL635',
}

wb = openpyxl.load_workbook(XLSX, data_only=True)

def fmt_date(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%Y-%m-%d')
    return str(v)

def match(eid, family, point, extra_desc=None):
    key = (eid, family, normalize(extra_desc if extra_desc is not None else point))
    vib_id = vib_lookup.get(key, '')
    if vib_id:
        return vib_id
    pos = derive_position_code(point)
    if pos:
        return poscode_lookup.get((eid, family, pos), '')
    return ''

def group_key(vib_id, eid, point):
    # Group by VIB ID when matched; otherwise by (Equipment ID, point text) so
    # unrelated unmatched points for the same equipment never get merged into
    # one row just because they share a blank VIB ID.
    return vib_id if vib_id else f'__UNMATCHED__|{eid}|{normalize(point)}'

# ---------- RMS ----------
ws = wb['📥 RMS DATA']
groups = defaultdict(list)   # (group_key, month) -> list of (date, eid, point, vib_id, axial, gear, horizontal, vertical)
for r in range(4, ws.max_row + 1):
    eid = ws.cell(row=r, column=3).value
    point = ws.cell(row=r, column=4).value
    if not eid and not point:
        continue
    eid = ID_CORRECTIONS.get(str(eid).strip(), str(eid).strip()) if eid else ''
    point = point if point else ''
    date = ws.cell(row=r, column=5).value
    axial = ws.cell(row=r, column=6).value
    gear = ws.cell(row=r, column=7).value
    horizontal = ws.cell(row=r, column=8).value
    vertical = ws.cell(row=r, column=9).value
    vib_id = match(eid, 'RMS', point)
    month = fmt_date(date)[:7]
    gk = group_key(vib_id, eid, point)
    sig = (fmt_date(date), axial, gear, horizontal, vertical)
    groups[(gk, month)].append((sig, eid, point, vib_id))

rms_rows = []
rms_overflow = []
for (gk, month), entries in groups.items():
    eid = entries[0][1]
    point = entries[0][2]
    vib_id = entries[0][3]
    # de-duplicate exact repeat rows (same date + same 4 values logged twice)
    seen = []
    for sig, *_ in entries:
        if sig not in seen:
            seen.append(sig)
    seen.sort(key=lambda s: s[0])  # sort by date
    row = [vib_id, eid, ('' if vib_id else point), month]
    for i in range(RMS_SLOTS):
        if i < len(seen):
            date, axial, gear, horizontal, vertical = seen[i]
            row += [date, axial, gear, horizontal, vertical]
        else:
            row += ['', '', '', '', '']
    rms_rows.append(row)
    if len(seen) > RMS_SLOTS:
        for date, axial, gear, horizontal, vertical in seen[RMS_SLOTS:]:
            rms_overflow.append([vib_id, eid, point, month, date, axial, gear, horizontal, vertical])

rms_rows.sort(key=lambda r: (r[1], r[3], r[0]))
print('RMS pivoted rows:', len(rms_rows), 'overflow readings beyond', RMS_SLOTS, 'slots:', len(rms_overflow))

# ---------- SPM + GS ----------
ws = wb['📥 SPM DATA']
spm_groups = defaultdict(list)
gs_groups = defaultdict(list)
for r in range(4, ws.max_row + 1):
    eid = ws.cell(row=r, column=3).value
    point = ws.cell(row=r, column=4).value
    if not eid and not point:
        continue
    eid = ID_CORRECTIONS.get(str(eid).strip(), str(eid).strip()) if eid else ''
    point = point if point else ''
    type_val = ws.cell(row=r, column=5).value
    date = ws.cell(row=r, column=6).value
    hdm = ws.cell(row=r, column=7).value
    hdc = ws.cell(row=r, column=8).value
    gs = ws.cell(row=r, column=9).value
    type_str = str(type_val).strip() if type_val is not None else ''
    desc = f'{point} ({type_str})' if type_str else point
    vib_id = match(eid, 'SPM', point, extra_desc=desc)
    month = fmt_date(date)[:7]
    gk = group_key(vib_id, eid, point)
    spm_groups[(gk, month)].append(((fmt_date(date), hdm, hdc), eid, point, vib_id))
    if gs is not None and gs != '':
        gs_groups[(gk, month)].append(((fmt_date(date), gs), eid, point, vib_id))

def build_pivot(groups_dict, n_slots, value_fields):
    out_rows = []
    overflow = []
    for (gk, month), entries in groups_dict.items():
        eid = entries[0][1]
        point = entries[0][2]
        vib_id = entries[0][3]
        seen = []
        for sig, *_ in entries:
            if sig not in seen:
                seen.append(sig)
        seen.sort(key=lambda s: s[0])
        row = [vib_id, eid, ('' if vib_id else point), month]
        for i in range(n_slots):
            if i < len(seen):
                row += list(seen[i])
            else:
                row += [''] * (1 + value_fields)
        out_rows.append(row)
        if len(seen) > n_slots:
            for extra in seen[n_slots:]:
                overflow.append([vib_id, eid, point, month] + list(extra))
    out_rows.sort(key=lambda r: (r[1], r[3], r[0]))
    return out_rows, overflow

spm_rows, spm_overflow = build_pivot(spm_groups, SPM_SLOTS, 2)
gs_rows, gs_overflow = build_pivot(gs_groups, SPM_SLOTS, 1)
print('SPM pivoted rows:', len(spm_rows), 'overflow:', len(spm_overflow))
print('GS pivoted rows:', len(gs_rows), 'overflow:', len(gs_overflow))

# ---------- write outputs ----------
def write_csv(path, header, rows):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(header)
        for i, row in enumerate(rows, start=1):
            w.writerow([i] + row)

rms_header = ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Month']
for i in range(1, RMS_SLOTS + 1):
    rms_header += [f'Date {i}', f'Axial {i} (mm/s)', f'Gear {i} (mm/s)', f'Horizontal {i} (mm/s)', f'Vertical {i} (mm/s)']
write_csv(OUT_DIR + 'rms-data-monthly.csv', rms_header, rms_rows)

spm_header = ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Month']
for i in range(1, SPM_SLOTS + 1):
    spm_header += [f'Date {i}', f'HDm {i} (dBsv)', f'HDc {i} (dBsv)']
write_csv(OUT_DIR + 'spm-data-monthly.csv', spm_header, spm_rows)

gs_header = ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Month']
for i in range(1, SPM_SLOTS + 1):
    gs_header += [f'Date {i}', f'Gs {i}']
write_csv(OUT_DIR + 'gs-data-monthly.csv', gs_header, gs_rows)

write_csv(OUT_DIR + 'rms-data-monthly-overflow.csv',
          ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Month', 'Date', 'Axial (mm/s)', 'Gear (mm/s)', 'Horizontal (mm/s)', 'Vertical (mm/s)'],
          rms_overflow)
write_csv(OUT_DIR + 'spm-data-monthly-overflow.csv',
          ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Month', 'Date', 'HDm (dBsv)', 'HDc (dBsv)'],
          spm_overflow)
write_csv(OUT_DIR + 'gs-data-monthly-overflow.csv',
          ['#', 'VIB ID', 'Equipment ID', 'Point (unmatched only)', 'Month', 'Date', 'Gs'],
          gs_overflow)

n_unmatched_rms = sum(1 for r in rms_rows if not r[0])
n_unmatched_spm = sum(1 for r in spm_rows if not r[0])
n_unmatched_gs = sum(1 for r in gs_rows if not r[0])
print('unmatched RMS pivot rows (blank VIB ID):', n_unmatched_rms)
print('unmatched SPM pivot rows (blank VIB ID):', n_unmatched_spm)
print('unmatched GS pivot rows (blank VIB ID):', n_unmatched_gs)
